# tools/mt5_symbol_inspector.py
"""
APEX MT5 Symbol Inspector
=========================
Standalone tool — connects to MT5 and discovers broker symbol names,
contract specs, tick values, and filling mode support for all configured tickers.

Usage:
    python tools/mt5_symbol_inspector.py
    python tools/mt5_symbol_inspector.py --login 12345 --password mypass --server ICMarkets-Demo

Outputs (to tools/output/):
    symbol_specs.xlsx  — 4-sheet Excel (Summary, Specs, Filling Modes, Recommendations)
    symbol_specs.yaml  — Machine-readable specs for config
    symbol_specs.csv   — Flat CSV for quick review

This script is FULLY STANDALONE — no APEX imports required.
Run it once after connecting to your broker to populate config/mt5_symbol_map.yaml.
"""
from __future__ import annotations

import argparse
import getpass
import json
import logging
import os
import sys
from dataclasses import dataclass, asdict, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import yaml

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "tools" / "output"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("mt5_inspector")

try:
    import MetaTrader5 as mt5
    MT5_AVAILABLE = True
except ImportError:
    MT5_AVAILABLE = False
    logger.error("MetaTrader5 not installed. Run: pip install MetaTrader5")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed — Excel export will be skipped.")

# ─── Constants ────────────────────────────────────────────────────────────────
FILL_FOK_BIT = 1
FILL_IOC_BIT = 2
FILL_RETURN_BIT = 4

SUFFIX_PATTERNS = [
    "",           # No suffix (some brokers)
    ".NYSE",      # Some ECN brokers
    ".NASDAQ",
    ".US",        # IC Markets, Pepperstone, FP Markets
    ".NAS",
    ".NYS",
    ".OTC",
    "m",          # Admiral Markets
    ".N",
    ".O",
    ".A",
]

US_TICKERS = [
    "AAPL", "MSFT", "GOOGL", "GOOG", "AMZN", "META", "NVDA", "TSLA",
    "JPM", "JNJ", "V", "UNH", "XOM", "WMT", "PG", "MA", "HD", "CVX",
    "MRK", "ABBV", "PEP", "KO", "COST", "BAC", "LLY", "AVGO", "ORCL",
    "ASML", "TXN", "CSCO", "NFLX", "AMD", "QCOM", "INTC", "IBM", "GE",
    "CAT", "DE", "HON", "MMM", "UPS", "FDX", "SBUX", "NKE", "MCD",
    "DIS", "CMCSA", "T", "VZ", "AMT", "NEE", "DUK", "SO",
]


@dataclass
class SymbolSpec:
    ticker: str
    mt5_symbol: str
    found: bool
    # Contract specs
    contract_size: float = 0.0
    tick_size: float = 0.0
    tick_value: float = 0.0
    point: float = 0.0
    currency_profit: str = ""
    currency_base: str = ""
    # Filling modes
    filling_mode_raw: int = 0
    supports_fok: bool = False
    supports_ioc: bool = False
    supports_return: bool = False
    recommended_filling: str = ""
    # Trade params
    volume_min: float = 0.0
    volume_max: float = 0.0
    volume_step: float = 0.0
    spread: float = 0.0
    stops_level: int = 0
    trade_mode: int = 0
    trade_mode_name: str = ""
    # Margin
    margin_initial: float = 0.0
    margin_maintenance: float = 0.0
    # Category
    category: str = ""
    exchange: str = ""
    description: str = ""
    # Discovery
    suffix_used: str = ""
    suffix_tried: int = 0
    error: str = ""


def connect_mt5(login: int, password: str, server: str, timeout_ms: int = 10000) -> bool:
    if not mt5.initialize():
        logger.error(f"MT5 initialize() failed: {mt5.last_error()}")
        return False

    ok = mt5.login(login=login, password=password, server=server, timeout=timeout_ms)
    if not ok:
        logger.error(f"MT5 login failed: {mt5.last_error()}")
        mt5.shutdown()
        return False

    info = mt5.account_info()
    logger.info(
        f"Connected: account={info.login}, server={info.server}, "
        f"equity={info.equity:.2f} {info.currency}"
    )
    return True


def find_symbol(ticker: str) -> Tuple[Optional[str], str, int]:
    """
    Try all suffix patterns to find the correct MT5 symbol for a ticker.
    Returns (mt5_symbol, suffix_used, tries_count) or (None, "", tries)
    """
    for i, suffix in enumerate(SUFFIX_PATTERNS, 1):
        candidate = f"{ticker}{suffix}"
        info = mt5.symbol_info(candidate)
        if info is not None:
            # Enable symbol in market watch
            mt5.symbol_select(candidate, True)
            return candidate, suffix, i
    return None, "", len(SUFFIX_PATTERNS)


def get_trade_mode_name(trade_mode: int) -> str:
    modes = {0: "FULL", 1: "LONG_ONLY", 2: "SHORT_ONLY", 4: "CLOSE_ONLY"}
    return modes.get(trade_mode, f"MODE_{trade_mode}")


def inspect_symbol(ticker: str) -> SymbolSpec:
    spec = SymbolSpec(ticker=ticker, mt5_symbol="", found=False)

    mt5_sym, suffix, tries = find_symbol(ticker)
    spec.suffix_tried = tries

    if mt5_sym is None:
        spec.error = f"Not found after {tries} suffix attempts"
        logger.warning(f"  {ticker}: NOT FOUND (tried {tries} suffixes)")
        return spec

    spec.mt5_symbol = mt5_sym
    spec.suffix_used = suffix
    spec.found = True

    try:
        info = mt5.symbol_info(mt5_sym)
        if info is None:
            spec.error = "symbol_info returned None"
            return spec

        spec.contract_size = float(info.trade_contract_size)
        spec.tick_size = float(info.trade_tick_size)
        spec.tick_value = float(info.trade_tick_value)
        spec.point = float(info.point)
        spec.currency_profit = str(info.currency_profit)
        spec.currency_base = str(info.currency_base)
        spec.filling_mode_raw = int(info.filling_mode)
        spec.supports_fok = bool(info.filling_mode & FILL_FOK_BIT)
        spec.supports_ioc = bool(info.filling_mode & FILL_IOC_BIT)
        spec.supports_return = bool(info.filling_mode & FILL_RETURN_BIT)
        spec.volume_min = float(info.volume_min)
        spec.volume_max = float(info.volume_max)
        spec.volume_step = float(info.volume_step)
        spec.stops_level = int(info.trade_stops_level)
        spec.trade_mode = int(info.trade_mode)
        spec.trade_mode_name = get_trade_mode_name(info.trade_mode)
        spec.margin_initial = float(info.margin_initial)
        spec.margin_maintenance = float(info.margin_maintenance)
        spec.description = str(info.description)

        # Spread from tick
        tick = mt5.symbol_info_tick(mt5_sym)
        if tick:
            spec.spread = round(float(tick.ask - tick.bid), 5)

        # Recommended filling mode
        if spec.supports_ioc:
            spec.recommended_filling = "IOC"
        elif spec.supports_fok:
            spec.recommended_filling = "FOK"
        elif spec.supports_return:
            spec.recommended_filling = "RETURN"
        else:
            spec.recommended_filling = "UNKNOWN"

        logger.info(
            f"  {ticker:8s} → {mt5_sym:15s} | "
            f"tick_val={spec.tick_value:.4f} | "
            f"filling={'FOK' if spec.supports_fok else ''}{'IOC' if spec.supports_ioc else ''}{'RETURN' if spec.supports_return else ''} | "
            f"rec={spec.recommended_filling}"
        )

    except Exception as exc:
        spec.error = str(exc)
        logger.error(f"  {ticker}: inspection error — {exc}")

    return spec


def inspect_all(tickers: List[str]) -> List[SymbolSpec]:
    specs = []
    logger.info(f"\nInspecting {len(tickers)} tickers...")
    for ticker in tickers:
        spec = inspect_symbol(ticker)
        specs.append(spec)
    found = sum(1 for s in specs if s.found)
    logger.info(f"\nDiscovery complete: {found}/{len(tickers)} symbols found.")
    return specs


# ─── Excel Export ─────────────────────────────────────────────────────────────
def _header_style(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="1F3864")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        bottom=Side(border_style="thin", color="FFFFFF"),
        right=Side(border_style="thin", color="DDDDDD"),
    )


def _data_style(ws, row: int, col: int, value: Any, alt: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    bg = "F2F2F2" if alt else "FFFFFF"
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        bottom=Side(border_style="hair", color="CCCCCC"),
        right=Side(border_style="hair", color="CCCCCC"),
    )


def export_excel(specs: List[SymbolSpec], path: Path) -> None:
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available — skipping Excel export.")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_summary_sheet(wb, specs)
    _build_specs_sheet(wb, specs)
    _build_filling_sheet(wb, specs)
    _build_recommendations_sheet(wb, specs)

    wb.save(path)
    logger.info(f"Excel saved: {path}")


def _build_summary_sheet(wb, specs: List[SymbolSpec]) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    headers = ["Ticker", "MT5 Symbol", "Found", "Suffix", "Rec Filling",
               "Tick Value", "Contract Size", "Spread", "Error"]
    widths = [10, 18, 8, 10, 12, 12, 14, 10, 30]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _header_style(ws, 1, c, h)
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.row_dimensions[1].height = 28

    for r, s in enumerate(specs, 2):
        alt = (r % 2 == 0)
        row_data = [
            s.ticker, s.mt5_symbol or "NOT FOUND",
            "✓" if s.found else "✗",
            f'"{s.suffix_used}"' if s.suffix_used else '(none)',
            s.recommended_filling,
            round(s.tick_value, 6) if s.found else "",
            s.contract_size if s.found else "",
            round(s.spread, 5) if s.found else "",
            s.error or "",
        ]
        for c, val in enumerate(row_data, 1):
            _data_style(ws, r, c, val, alt)
            cell = ws.cell(row=r, column=c)
            if c == 3:
                cell.font = Font(
                    color="2E7D32" if s.found else "C62828",
                    bold=True
                )
            if c == 5 and s.found:
                color = "1565C0" if s.recommended_filling == "IOC" else "4A148C"
                cell.font = Font(color=color, bold=True)


def _build_specs_sheet(wb, specs: List[SymbolSpec]) -> None:
    ws = wb.create_sheet("Contract Specs")
    ws.sheet_view.showGridLines = False

    headers = ["Ticker", "MT5 Symbol", "Contract Size", "Tick Size",
               "Tick Value", "Point", "Vol Min", "Vol Max", "Vol Step",
               "Stops Level", "Margin Initial", "Currency Profit", "Description"]
    widths = [10, 18, 14, 12, 12, 10, 10, 12, 10, 12, 15, 16, 35]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _header_style(ws, 1, c, h)
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.row_dimensions[1].height = 28

    for r, s in enumerate([x for x in specs if x.found], 2):
        alt = (r % 2 == 0)
        row = [s.ticker, s.mt5_symbol, s.contract_size, s.tick_size,
               round(s.tick_value, 6), s.point, s.volume_min, s.volume_max,
               s.volume_step, s.stops_level, s.margin_initial,
               s.currency_profit, s.description]
        for c, val in enumerate(row, 1):
            _data_style(ws, r, c, val, alt)


def _build_filling_sheet(wb, specs: List[SymbolSpec]) -> None:
    ws = wb.create_sheet("Filling Modes")
    ws.sheet_view.showGridLines = False

    headers = ["Ticker", "MT5 Symbol", "Raw Bitmask",
               "FOK (bit 1)", "IOC (bit 2)", "RETURN (bit 4)", "Recommended"]
    widths = [10, 18, 12, 12, 12, 14, 14]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _header_style(ws, 1, c, h)
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.row_dimensions[1].height = 28

    for r, s in enumerate([x for x in specs if x.found], 2):
        alt = (r % 2 == 0)
        row = [
            s.ticker, s.mt5_symbol, s.filling_mode_raw,
            "✓" if s.supports_fok else "✗",
            "✓" if s.supports_ioc else "✗",
            "✓" if s.supports_return else "✗",
            s.recommended_filling,
        ]
        for c, val in enumerate(row, 1):
            cell_obj = ws.cell(row=r, column=c)
            cell_obj.value = val
            bg = "F2F2F2" if alt else "FFFFFF"
            cell_obj.fill = PatternFill("solid", fgColor=bg)
            cell_obj.alignment = Alignment(horizontal="center", vertical="center")
            cell_obj.border = Border(
                bottom=Side(border_style="hair", color="CCCCCC"),
                right=Side(border_style="hair", color="CCCCCC"),
            )
            if c in (4, 5, 6):
                cell_obj.font = Font(
                    color="2E7D32" if val == "✓" else "C62828",
                    bold=True,
                )
            if c == 7:
                color_map = {"IOC": "1565C0", "FOK": "4A148C", "RETURN": "E65100"}
                cell_obj.font = Font(color=color_map.get(val, "000000"), bold=True)


def _build_recommendations_sheet(wb, specs: List[SymbolSpec]) -> None:
    ws = wb.create_sheet("YAML Recommendations")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 60

    ws.cell(row=1, column=1, value="YAML Config — copy to config/mt5_symbol_map.yaml").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value="")

    lines = ["symbols:"]
    for s in specs:
        if s.found:
            lines.append(f'  {s.ticker}: "{s.mt5_symbol}"')
        else:
            lines.append(f'  {s.ticker}: ""  # NOT FOUND — check broker symbol list')

    lines.append("")
    lines.append("filling_overrides:  # Auto-detected recommendations:")
    for s in [x for x in specs if x.found and x.recommended_filling]:
        comment = ""
        if not s.supports_ioc and not s.supports_fok:
            comment = "  # WARNING: only RETURN supported"
        lines.append(f"  {s.mt5_symbol}: \"{s.recommended_filling}\"{comment}")

    for i, line in enumerate(lines, 3):
        ws.cell(row=i, column=1, value=line).font = Font(name="Courier New", size=10)

    ws.column_dimensions["A"].width = 70


# ─── YAML / CSV Export ────────────────────────────────────────────────────────
def export_yaml(specs: List[SymbolSpec], path: Path) -> None:
    data = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "symbols": {},
        "filling_overrides": {},
        "not_found": [],
    }
    for s in specs:
        if s.found:
            data["symbols"][s.ticker] = s.mt5_symbol
            data["filling_overrides"][s.mt5_symbol] = s.recommended_filling
        else:
            data["not_found"].append(s.ticker)

    path.write_text(yaml.dump(data, default_flow_style=False, sort_keys=False), encoding="utf-8")
    logger.info(f"YAML saved: {path}")


def export_csv(specs: List[SymbolSpec], path: Path) -> None:
    rows = [asdict(s) for s in specs]
    df = pd.DataFrame(rows)
    df.to_csv(path, index=False, encoding="utf-8")
    logger.info(f"CSV saved: {path}")


# ─── Main ─────────────────────────────────────────────────────────────────────
def get_credentials(args: argparse.Namespace) -> Tuple[int, str, str]:
    login = args.login or int(os.environ.get("MT5_LOGIN", "0") or "0")
    password = args.password or os.environ.get("MT5_PASSWORD", "")
    server = args.server or os.environ.get("MT5_SERVER", "")

    if not login:
        login = int(input("MT5 Login (account number): ").strip())
    if not password:
        password = getpass.getpass("MT5 Password: ")
    if not server:
        server = input("MT5 Server: ").strip()

    return login, password, server


def main() -> None:
    parser = argparse.ArgumentParser(description="APEX MT5 Symbol Inspector")
    parser.add_argument("--login", type=int, default=0)
    parser.add_argument("--password", type=str, default="")
    parser.add_argument("--server", type=str, default="")
    parser.add_argument(
        "--tickers",
        type=str,
        default="",
        help="Comma-separated ticker list (default: built-in US_TICKERS)",
    )
    parser.add_argument(
        "--extra",
        type=str,
        default="",
        help="Additional comma-separated tickers to append",
    )
    args = parser.parse_args()

    login, password, server = get_credentials(args)

    if not connect_mt5(login, password, server):
        sys.exit(1)

    try:
        tickers = (
            [t.strip() for t in args.tickers.split(",") if t.strip()]
            if args.tickers
            else list(US_TICKERS)
        )
        if args.extra:
            extras = [t.strip() for t in args.extra.split(",") if t.strip()]
            tickers = list(dict.fromkeys(tickers + extras))

        specs = inspect_all(tickers)

        OUT_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

        xlsx_path = OUT_DIR / f"symbol_specs_{ts}.xlsx"
        yaml_path = OUT_DIR / f"symbol_specs_{ts}.yaml"
        csv_path = OUT_DIR / f"symbol_specs_{ts}.csv"

        # Also write latest (no timestamp) for easy reference
        latest_xlsx = OUT_DIR / "symbol_specs_latest.xlsx"
        latest_yaml = OUT_DIR / "symbol_specs_latest.yaml"
        latest_csv = OUT_DIR / "symbol_specs_latest.csv"

        export_excel(specs, xlsx_path)
        export_yaml(specs, yaml_path)
        export_csv(specs, csv_path)

        export_excel(specs, latest_xlsx)
        export_yaml(specs, latest_yaml)
        export_csv(specs, latest_csv)

        # Print summary
        found = [s for s in specs if s.found]
        not_found = [s for s in specs if not s.found]

        print(f"\n{'─' * 60}")
        print(f"  APEX MT5 Symbol Inspector — Results")
        print(f"{'─' * 60}")
        print(f"  Total tickers  : {len(specs)}")
        print(f"  Found          : {len(found)}")
        print(f"  Not found      : {len(not_found)}")
        print(f"\n  Filling mode distribution:")
        ioc_count = sum(1 for s in found if s.recommended_filling == "IOC")
        fok_count = sum(1 for s in found if s.recommended_filling == "FOK")
        ret_count = sum(1 for s in found if s.recommended_filling == "RETURN")
        print(f"    IOC    : {ioc_count}")
        print(f"    FOK    : {fok_count}")
        print(f"    RETURN : {ret_count}")

        if not_found:
            print(f"\n  Not found: {', '.join(s.ticker for s in not_found)}")

        print(f"\n  Outputs:")
        print(f"    {xlsx_path}")
        print(f"    {yaml_path}")
        print(f"    {csv_path}")
        print(f"{'─' * 60}\n")
        print("  → Copy symbol names from YAML output into config/mt5_symbol_map.yaml")
        print()

    finally:
        mt5.shutdown()
        logger.info("MT5 disconnected.")


if __name__ == "__main__":
    main()
